import os
import re
import openpyxl
from github import Github
from github.ContentFile import ContentFile
from typing import List, Tuple, Optional
from dotenv import load_dotenv

# Load environment variables from .env file
load_dotenv()

# Constants
PASSWORD_REGEX = r"(password|passwd|pwd|secret|token|key)\s*[:=]\s*['\"]?([^\s'\"&;]+)['\"]?"
EXCEL_FILE = "hardcoded_passwords.xlsx"
SHEET_NAME = "Passwords"
START_ROW = 1
START_COL = 1

def check_file_for_passwords(file_content: str, file_path: str) -> List[Tuple[str, int, str]]:
    """
    Checks a single file's content for potential hardcoded passwords.

    Args:
        file_content (str): The content of the file to check.
        file_path (str): The path of the file (for reporting).

    Returns:
        List[Tuple[str, int, str]]: A list of tuples, where each tuple contains:
            - The matched string (the potential password).
            - The line number where the match was found.
            - The full line of code where the match was found.
            Returns an empty list if no matches are found.
    """
    matches = []
    lines = file_content.splitlines()
    for i, line in enumerate(lines):
        match = re.search(PASSWORD_REGEX, line, re.IGNORECASE)
        if match:
            password = match.group(2)  # Extract the actual password-like string
            matches.append((password, i + 1, line.strip()))  # Store line number and full line
    return matches

def process_repository(repo, processed_files: set) -> List[Tuple[str, str, int, str]]:
    """
    Recursively processes a GitHub repository to find potential hardcoded passwords.  Handles submodules
    and prevents infinite recursion.

    Args:
        repo (github.Repository.Repository): The GitHub repository object to process.
        processed_files (set): A set to keep track of processed file paths (to avoid duplicates).

    Returns:
        List[Tuple[str, str, int, str]]: A list of tuples, where each tuple represents a found
        hardcoded password: (file_path, matched_string, line_number, full_line).
    """
    results = []
    contents = repo.get_contents("")  # Get the contents of the root directory
    while contents:
        content = contents.pop(0)
        if content.type == "file":
            if content.path in processed_files:
                continue  # Skip files already processed
            processed_files.add(content.path)  # Mark file as processed
            try:
                file_content = content.decoded_content.decode("utf-8", errors="ignore")
                matches = check_file_for_passwords(file_content, content.path)
                for password, line_number, full_line in matches:
                    results.append((content.path, password, line_number, full_line))
            except Exception as e:
                print(f"Error reading file {content.path}: {e}")
        elif content.type == "dir":
            contents.extend(repo.get_contents(content.path))
        elif content.type == "submodule":
            try:
                # IMPORTANT:  Handle submodules to avoid missing secrets!
                submodule = repo.get_submodule(content.path)
                #  Get the submodule's repository object.  This is CRUCIAL.
                sub_repo = submodule.module() # Changed from submodule.repository to submodule.module()
                if sub_repo:
                    results.extend(process_repository(sub_repo, processed_files)) # Recursive call for submodules
                else:
                    print(f"Warning: Could not get repository object for submodule: {content.path}")

            except Exception as e:
                print(f"Error processing submodule {content.path}: {e}")
        else:
            print(f"Unknown content type: {content.type} in {content.path}")
    return results

def write_to_excel(results: List[Tuple[str, str, int, str]]):
    """
    Writes the results to an Excel file.

    Args:
        results (List[Tuple[str, str, int, str]]): A list of tuples, where each tuple represents a
            found hardcoded password: (file_path, matched_string, line_number, full_line).
    """
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = SHEET_NAME

    # Write headers
    headers = ["File Path", "Matched String", "Line Number", "Full Line"]
    for col, header in enumerate(headers, start=START_COL):
        sheet.cell(row=START_ROW, column=col, value=header)

    # Write data
    for row, (file_path, matched_string, line_number, full_line) in enumerate(results, start=START_ROW + 1):
        sheet.cell(row=row, column=START_COL, value=file_path)
        sheet.cell(row=row, column=START_COL + 1, value=matched_string)
        sheet.cell(row=row, column=START_COL + 2, value=line_number)
        sheet.cell(row=row, column=START_COL + 3, value=full_line)

    workbook.save(EXCEL_FILE)
    print(f"Results written to {EXCEL_FILE}")

def main():
    """
    Main function to drive the script.
    - Gets the GitHub token and repository name.
    - Initializes the GitHub connection.
    - Processes the repository.
    - Writes the results to an Excel file.
    """
    github_token = os.environ.get("GITHUB_TOKEN")
    if not github_token:
        print("Error: GITHUB_TOKEN environment variable not set.")
        print("Please set the GITHUB_TOKEN environment variable with your GitHub personal access token.")
        print("You can create a personal access token here: https://github.com/settings/tokens")
        return

    repo_name = os.environ.get("GITHUB_REPO")
    if not repo_name:
        print("Error: GITHUB_REPO environment variable not set.")
        print("Please set the GITHUB_REPO environment variable with the name of the repository to scan (e.g., 'your_username/your_repo_name').")
        return

    try:
        # Initialize GitHub connection using a Personal Access Token
        g = Github(github_token)
        repo = g.get_repo(repo_name)  # Use the full repository name
        print(f"Scanning repository: {repo.full_name}")  # Print the full name
    except Exception as e:
        print(f"Error connecting to GitHub or accessing the repository: {e}")
        print("Please check your GITHUB_TOKEN and GITHUB_REPO environment variables.")
        return

    processed_files = set() # Keep track of processed files.
    results = process_repository(repo, processed_files)
    if results:
        write_to_excel(results)
    else:
        print("No potential hardcoded passwords found.")

if __name__ == "__main__":
    main()

